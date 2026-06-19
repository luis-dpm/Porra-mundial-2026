#!/usr/bin/env python3
"""
Actualiza data.js con los resultados reales del Mundial 2026.

Fuente principal de resultados: el propio Excel (predicciones.xlsx),
que el administrador de la porra mantiene actualizado a mano.
La API de football-data.org se usa solo como complemento, para
rellenar resultados que falten en el Excel — nunca para sobreescribir
lo que ya hay en el Excel.

Las fechas de los partidos se toman del calendario oficial en horario
de España peninsular (match_dates_es.py), no del Excel ni de la API,
porque ambos pueden traer la fecha en huso horario de EE.UU./México.

Variables de entorno requeridas:
  FOOTBALL_DATA_API_KEY - clave de football-data.org (opcional; si falta
                            o falla, se usa solo lo que haya en el Excel)
"""
import os
import re
import json
import sys
from datetime import date
from collections import defaultdict

from openpyxl import load_workbook

try:
    import requests
except ImportError:
    requests = None

from match_dates_es import MATCH_DATES_ES

EXCEL_PATH = "source/predicciones.xlsx"
OUTPUT_PATH = "data.js"
COMPETITION_CODE = "WC"
API_BASE = "https://api.football-data.org/v4"

# Mapeo nombre del Excel (español) -> nombre usado por football-data.org (inglés)
TEAM_NAME_MAP = {
    "Alemania": "Germany", "Arabia Saudita": "Saudi Arabia", "Argelia": "Algeria",
    "Argentina": "Argentina", "Australia": "Australia", "Austria": "Austria",
    "Bosnia y Herzegovina": "Bosnia and Herzegovina", "Brasil": "Brazil",
    "Bélgica": "Belgium", "Cabo Verde": "Cape Verde", "Canadá": "Canada",
    "Catar": "Qatar", "Colombia": "Colombia", "Corea del Sur": "South Korea",
    "Costa de Marfil": "Ivory Coast", "Croacia": "Croatia", "Curazao": "Curacao",
    "Ecuador": "Ecuador", "Egipto": "Egypt", "Escocia": "Scotland",
    "España": "Spain", "Estados Unidos": "United States", "Francia": "France",
    "Ghana": "Ghana", "Haití": "Haiti", "Inglaterra": "England", "Irak": "Iraq",
    "Irán": "Iran", "Japón": "Japan", "Jordania": "Jordan", "Marruecos": "Morocco",
    "México": "Mexico", "Noruega": "Norway", "Nueva Zelanda": "New Zealand",
    "Panamá": "Panama", "Paraguay": "Paraguay", "Países Bajos": "Netherlands",
    "Portugal": "Portugal", "RD Congo": "DR Congo", "República Checa": "Czech Republic",
    "Senegal": "Senegal", "Sudáfrica": "South Africa", "Suecia": "Sweden",
    "Suiza": "Switzerland", "Turquía": "Turkey", "Túnez": "Tunisia",
    "Uruguay": "Uruguay", "Uzbekistán": "Uzbekistan",
}
TEAM_NAME_MAP_REV = {v: k for k, v in TEAM_NAME_MAP.items()}

# Nombres EXACTOS tal y como aparecen en la fila 5 del Excel (columna ADMIN).
# Si cambian de nombre en el Excel, hay que actualizarlos aquí también.
PLAYER_COLUMNS = {
    "LUIS DPM": 19,
    "EL MATO": 22,
    "IVÁN DELGADO": 25,
    "ADRIÁN": 28,
    "JUAN": 31,
    "CARLOS": 34,
    "SU FLORENTINEZA": 37,
}


def fetch_real_results(api_key):
    """Llama a football-data.org. Si falla por cualquier motivo, devuelve {}
    y el script sigue funcionando solo con lo que haya en el Excel."""
    if not api_key or not requests:
        return {}
    try:
        headers = {"X-Auth-Token": api_key}
        url = f"{API_BASE}/competitions/{COMPETITION_CODE}/matches"
        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        print(f"AVISO: no se pudo consultar la API ({e}). Se usará solo el Excel.", file=sys.stderr)
        return {}

    results = {}
    for m in data.get("matches", []):
        if m["status"] != "FINISHED":
            continue
        home_en = m["homeTeam"]["name"]
        away_en = m["awayTeam"]["name"]
        home_es = TEAM_NAME_MAP_REV.get(home_en)
        away_es = TEAM_NAME_MAP_REV.get(away_en)
        if not home_es or not away_es:
            continue
        hg = m["score"]["fullTime"]["home"]
        ag = m["score"]["fullTime"]["away"]
        if hg is None or ag is None:
            continue
        results[(home_es, away_es)] = (hg, ag)
    return results


def read_excel_data():
    """Lee predicciones, resultados YA introducidos a mano y posiciones de grupo."""
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["ADMIN"]

    matches = []
    for row_idx in range(6, 270):
        match_name = ws.cell(row=row_idx, column=11).value
        if not match_name or match_name in ["-", "Casa-Fuera"]:
            continue
        if str(match_name).startswith("W") or str(match_name).startswith("L"):
            continue
        phase_cell = ws.cell(row=row_idx, column=10).value
        if not (phase_cell and re.match(r"^[A-L][123]$", str(phase_cell))):
            continue

        excel_actual = ws.cell(row=row_idx, column=13).value
        excel_actual = str(excel_actual).strip() if excel_actual and "|" in str(excel_actual) else None

        preds = {}
        for player, col in PLAYER_COLUMNS.items():
            pred = ws.cell(row=row_idx, column=col).value
            preds[player] = str(pred).strip() if pred and "|" in str(pred) else None

        matches.append({
            "match": match_name,
            "group": str(phase_cell)[0],
            "excel_actual": excel_actual,
            "predictions": preds,
        })

    group_positions = {}
    row = 80
    for g in "ABCDEFGHIJKL":
        group_positions[g] = {}
        for pos in [1, 2, 3, 4]:
            preds = {p: ws.cell(row=row, column=c).value for p, c in PLAYER_COLUMNS.items()}
            group_positions[g][pos] = preds
            row += 1

    return matches, group_positions


def score_match(pred_str, actual_str):
    if not pred_str or not actual_str:
        return {"pts": 0, "sign": False, "diff": False, "exact": False}
    psign, pscore = pred_str.split("|")
    ph, pa = map(int, pscore.split("-"))
    asign, ascore = actual_str.split("|")
    ah, aa = map(int, ascore.split("-"))
    pts = 0
    sign_ok = psign == asign
    diff_ok = exact_ok = False
    if sign_ok:
        pts += 1
        if abs(ph - pa) == abs(ah - aa):
            diff_ok = True
            pts += 1
        if ph == ah and pa == aa:
            exact_ok = True
            pts += 2
    return {"pts": pts, "sign": sign_ok, "diff": diff_ok, "exact": exact_ok}


def sign_from_score(h, a):
    return "1" if h > a else ("X" if h == a else "2")


def build_dataset(api_key):
    matches, group_positions = read_excel_data()
    api_results = fetch_real_results(api_key)
    players = list(PLAYER_COLUMNS.keys())

    processed = []
    group_team_stats = defaultdict(lambda: defaultdict(lambda: {"pj": 0, "g": 0, "e": 0, "p": 0, "gf": 0, "gc": 0}))
    matches_missing_date = []

    for m in matches:
        home, away = m["match"].split("-", 1)

        official_date = MATCH_DATES_ES.get(m["match"])
        if not official_date:
            matches_missing_date.append(m["match"])

        # Prioridad: 1) Excel (verificado a mano), 2) API si el Excel no lo tiene
        actual = None
        if m["excel_actual"]:
            actual = m["excel_actual"]
        else:
            api_res = api_results.get((home, away))
            if api_res:
                hg, ag = api_res
                actual = f"{sign_from_score(hg, ag)}|{hg}-{ag}"

        m2 = {
            "match": m["match"],
            "group": m["group"],
            "date": official_date,
            "actual": actual,
            "predictions": m["predictions"],
        }

        if actual:
            asign, ascore = actual.split("|")
            ah, ag = map(int, ascore.split("-"))

            breakdown = {}
            for p in players:
                breakdown[p] = score_match(m["predictions"].get(p), actual)
            m2["breakdown"] = breakdown

            t_h = group_team_stats[m["group"]][home]
            t_a = group_team_stats[m["group"]][away]
            t_h["pj"] += 1
            t_a["pj"] += 1
            t_h["gf"] += ah
            t_h["gc"] += ag
            t_a["gf"] += ag
            t_a["gc"] += ah
            if ah > ag:
                t_h["g"] += 1
                t_a["p"] += 1
            elif ah < ag:
                t_a["g"] += 1
                t_h["p"] += 1
            else:
                t_h["e"] += 1
                t_a["e"] += 1

        processed.append(m2)

    if matches_missing_date:
        print(f"AVISO: sin fecha oficial para: {matches_missing_date}", file=sys.stderr)

    group_standings_real = {}
    for g, teams in group_team_stats.items():
        rows = []
        for team, s in teams.items():
            pts = s["g"] * 3 + s["e"]
            gd = s["gf"] - s["gc"]
            rows.append({"team": team, **s, "gd": gd, "pts": pts})
        rows.sort(key=lambda r: (-r["pts"], -r["gd"], -r["gf"]))
        group_standings_real[g] = rows

    by_date = defaultdict(list)
    for m in processed:
        if m.get("date") and m.get("actual"):
            by_date[m["date"]].append(m)
    dates = sorted(by_date.keys())

    daily = {p: {} for p in players}
    cum = {p: {} for p in players}
    running = {p: 0 for p in players}
    for d in dates:
        for p in players:
            day_pts = sum(m["breakdown"][p]["pts"] for m in by_date[d])
            daily[p][d] = day_pts
            running[p] += day_pts
            cum[p][d] = running[p]

    standings = sorted(
        ({"player": p, "points": running[p]} for p in players),
        key=lambda x: -x["points"],
    )
    for i, s in enumerate(standings):
        s["rank"] = i + 1

    positions_by_day = {p: [] for p in players}
    for d in dates:
        sorted_p = sorted(players, key=lambda p: -cum[p][d])
        pos, prev, rank = {}, None, 0
        for i, p in enumerate(sorted_p):
            if cum[p][d] != prev:
                rank = i + 1
            pos[p] = rank
            prev = cum[p][d]
        for p in players:
            positions_by_day[p].append(pos[p])

    return {
        "matches": processed,
        "group_positions": group_positions,
        "group_standings_real": group_standings_real,
        "players": players,
        "dates": dates,
        "daily_points": daily,
        "cumulative_points": cum,
        "positions_by_day": positions_by_day,
        "standings": standings,
        "last_updated": date.today().isoformat(),
    }


def main():
    api_key = os.environ.get("FOOTBALL_DATA_API_KEY")
    dataset = build_dataset(api_key)

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write("const PORRA_DATA = ")
        json.dump(dataset, f, ensure_ascii=False)
        f.write(";")

    played = sum(1 for m in dataset["matches"] if m["actual"])
    print(f"OK: data.js actualizado. {played}/{len(dataset['matches'])} partidos jugados. "
          f"Líder: {dataset['standings'][0]['player']} ({dataset['standings'][0]['points']} pts)")


if __name__ == "__main__":
    main()
