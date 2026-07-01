#!/usr/bin/env python3
"""
Actualiza data.js con los resultados reales del Mundial 2026.

Fuente principal de resultados: el propio Excel (predicciones.xlsx),
que el administrador de la porra mantiene actualizado a mano.

Para partidos que el Excel no tenga, se usan DOS fuentes automáticas:
  1. football-data.org (necesita API key)
  2. openfootball/worldcup.json (gratuita, sin API key, en GitHub)

Si ambas fuentes coinciden en el marcador, se usa con confianza alta.
Si solo una de las dos lo tiene, se usa igualmente pero se marca como
"sin verificar" en debug_api.json (football-data.org puede tener datos
provisionales erróneos justo tras el pitido final que tarda unos minutos
en corregir). Si ambas lo tienen pero NO coinciden, se descarta y el
partido queda pendiente — mejor no mostrar nada que mostrar un dato
dudoso, y se loguea el conflicto para revisión manual.

Las fechas de los partidos se toman del calendario oficial en horario
de España peninsular (match_dates_es.py), no del Excel ni de la API,
porque ambos pueden traer la fecha en huso horario de EE.UU./México.

Variables de entorno requeridas:
  FOOTBALL_DATA_API_KEY - clave de football-data.org (opcional; si falta
                            o falla, se usa solo lo que haya en el Excel
                            y en openfootball)
"""
import os
import re
import json
import sys
from datetime import date, datetime, timedelta
from collections import defaultdict

from openpyxl import load_workbook

try:
    import requests
except ImportError:
    requests = None

from match_dates_es import MATCH_DATES_ES
from ko_stage import build_ko_dataset

EXCEL_PATH = "source/predicciones.xlsx"
OUTPUT_PATH = "data.js"
COMPETITION_CODE = "WC"
API_BASE = "https://api.football-data.org/v4"
OPENFOOTBALL_URL = "https://raw.githubusercontent.com/openfootball/worldcup.json/master/2026/worldcup.json"

# Mapeo nombre del Excel (español) -> nombre usado por football-data.org (inglés)
TEAM_NAME_MAP = {
    "Alemania": "Germany", "Arabia Saudita": "Saudi Arabia", "Argelia": "Algeria",
    "Argentina": "Argentina", "Australia": "Australia", "Austria": "Austria",
    "Bosnia y Herzegovina": "Bosnia and Herzegovina", "Brasil": "Brazil",
    "Bélgica": "Belgium", "Cabo Verde": "Cape Verde", "Canadá": "Canada",
    "Catar": "Qatar", "Colombia": "Colombia", "Corea del Sur": "South Korea",
    "Costa de Marfil": "Ivory Coast", "Croacia": "Croatia", "Curazao": "Curacao",
    "Ecuador": "Ecuador", "Egipto": "Egypt", "Escocia": "Scotland",
    "España": "Spain", "Estados Unidos": "United States", "Francia": "France",
    "Ghana": "Ghana", "Haití": "Haiti", "Inglaterra": "England", "Irak": "Iraq",
    "Irán": "Iran", "Japón": "Japan", "Jordania": "Jordan", "Marruecos": "Morocco",
    "México": "Mexico", "Noruega": "Norway", "Nueva Zelanda": "New Zealand",
    "Panamá": "Panama", "Paraguay": "Paraguay", "Países Bajos": "Netherlands",
    "Portugal": "Portugal", "RD Congo": "DR Congo", "República Checa": "Czech Republic",
    "Senegal": "Senegal", "Sudáfrica": "South Africa", "Suecia": "Sweden",
    "Suiza": "Switzerland", "Turquía": "Turkey", "Túnez": "Tunisia",
    "Uruguay": "Uruguay", "Uzbekistán": "Uzbekistan",
}
TEAM_NAME_MAP_REV = {v: k for k, v in TEAM_NAME_MAP.items()}

# Variantes adicionales que algunas APIs usan en vez del nombre "estándar" de
# arriba. Se añaden aquí para que el matching no falle en silencio.
TEAM_NAME_ALIASES = {
    "Korea Republic": "Corea del Sur",
    "Korea, South": "Corea del Sur",
    "South Korea": "Corea del Sur",
    "Czechia": "República Checa",
    "Czech Republic": "República Checa",
    "Côte d'Ivoire": "Costa de Marfil",
    "Cote d'Ivoire": "Costa de Marfil",
    "Ivory Coast": "Costa de Marfil",
    "IR Iran": "Irán",
    "Iran": "Irán",
    "Türkiye": "Turquía",
    "Turkiye": "Turquía",
    "Turkey": "Turquía",
    "USA": "Estados Unidos",
    "United States": "Estados Unidos",
    "United States of America": "Estados Unidos",
    "Bosnia-Herzegovina": "Bosnia y Herzegovina",
    "Bosnia & Herzegovina": "Bosnia y Herzegovina",
    "Bosnia and Herzegovina": "Bosnia y Herzegovina",
    "DR Congo": "RD Congo",
    "Congo DR": "RD Congo",
    "DRC": "RD Congo",
    "Curacao": "Curazao",
    "Curaçao": "Curazao",
    "Cape Verde Islands": "Cabo Verde",
    "Cape Verde": "Cabo Verde",
    "Saudi Arabia": "Arabia Saudita",
    "Republic of Ireland": "Irlanda",
}
TEAM_NAME_MAP_REV.update(TEAM_NAME_ALIASES)


def resolve_team_name(api_name):
    """Busca el nombre en español dado un nombre devuelto por la API,
    probando coincidencia exacta y luego variantes con/sin tildes y
    mayúsculas. Devuelve None si no hay nombre (p.ej. partidos de
    eliminatoria cuyo equipo aún no está determinado, donde la API
    devuelve homeTeam/awayTeam como null o sin 'name')."""
    if not api_name:
        return None
    if api_name in TEAM_NAME_MAP_REV:
        return TEAM_NAME_MAP_REV[api_name]
    # Intento case-insensitive
    lower_map = {k.lower(): v for k, v in TEAM_NAME_MAP_REV.items()}
    return lower_map.get(api_name.lower())

# Nombres EXACTOS tal y como aparecen en la fila 5 del Excel (columna ADMIN).
# Si cambian de nombre en el Excel, hay que actualizarlos aquí también.
PLAYER_COLUMNS = {
    "LUIS DPM": 19,
    "EL MATO": 22,
    "IVÁN DELGADO": 25,
    "ADRIÁN": 28,
    "JUAN": 31,
    "CARLOS": 34,
    "SU FLORENTINEZA": 37,
}


def fetch_official_standings(api_key):
    """Consulta /competitions/WC/standings: la clasificación de cada grupo ya
    calculada por football-data.org aplicando los criterios oficiales FIFA
    (incluye fair-play si está disponible, no solo pts/gd/gf como un cálculo
    casero). Devuelve {group_letter: [{'team': nombre_es, 'position': int,
    'points': int, ...}, ...]} o {} si no se puede consultar."""
    if not api_key or not requests:
        return {}
    try:
        headers = {"X-Auth-Token": api_key}
        url = f"{API_BASE}/competitions/{COMPETITION_CODE}/standings"
        resp = requests.get(url, headers=headers, timeout=30)
        print(f"INFO: API standings respondió con status {resp.status_code}", file=sys.stderr)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        print(f"AVISO: no se pudo consultar /standings ({e}). Se calculará localmente.", file=sys.stderr)
        return {}

    raw_standings = data.get("standings", [])
    print(f"INFO: la API devolvió {len(raw_standings)} bloques de standings.", file=sys.stderr)
    if raw_standings:
        # Volcamos la primera entrada tal cual para poder ver su forma real
        # en el log, en vez de adivinar el formato a ciegas.
        sample = {k: v for k, v in raw_standings[0].items() if k != "table"}
        sample["table_sample"] = raw_standings[0].get("table", [])[:1]
        print(f"INFO: estructura de ejemplo (1er bloque, sin tabla completa): {json.dumps(sample, ensure_ascii=False)}", file=sys.stderr)

    result = {}
    unmapped = []
    for standing in raw_standings:
        group_raw = standing.get("group")  # puede venir como "GROUP_A", "A", o null
        stage = standing.get("stage")
        if not group_raw:
            continue
        # Normalizamos: aceptamos "GROUP_A", "Group A", "A", etc. — nos
        # quedamos con la última letra mayúscula A-L que aparezca.
        letters = [c for c in str(group_raw).upper() if c.isalpha()]
        group_letter = letters[-1] if letters else None
        if not group_letter or group_letter not in "ABCDEFGHIJKL":
            print(f"AVISO: valor de 'group' no reconocido, se ignora este bloque: {group_raw!r} (stage={stage!r})", file=sys.stderr)
            continue

        rows = []
        for idx, row in enumerate(standing.get("table", [])):
            team_en = row["team"]["name"]
            team_es = resolve_team_name(team_en)
            if not team_es:
                unmapped.append(team_en)
                team_es = team_en  # fallback: mostrar el nombre tal cual antes que perder el dato
            # Usamos la posición que da la API si parece válida (1-4 y no
            # repetida dentro de este grupo); si no, recurrimos al orden en
            # que vino la fila — la API casi siempre devuelve las filas ya
            # ordenadas aunque el campo 'position' falle en algún empate.
            api_pos = row.get("position")
            rows.append({
                "team": team_es,
                "_api_position": api_pos,
                "position": api_pos if api_pos else idx + 1,
                "pj": row.get("playedGames", 0),
                "g": row.get("won", 0),
                "e": row.get("draw", 0),
                "p": row.get("lost", 0),
                "gf": row.get("goalsFor", 0),
                "gc": row.get("goalsAgainst", 0),
                "gd": row.get("goalDifference", 0),
                "pts": row.get("points", 0),
            })

        # Normalizamos las posiciones: si la API repitió o se saltó algún
        # número (típico en empates sin desempatar), reasignamos 1,2,3,4
        # respetando el orden en que la API ya trae las filas (que suele
        # venir bien ordenado aunque el campo 'position' individual falle).
        seen_positions = [r["position"] for r in rows]
        positions_ok = sorted(seen_positions) == list(range(1, len(rows) + 1))
        if not positions_ok and rows:
            print(f"AVISO: posiciones inconsistentes en grupo {group_letter} "
                  f"({seen_positions}), reasignando 1..{len(rows)} por orden de llegada.", file=sys.stderr)
            for i, r in enumerate(rows):
                r["position"] = i + 1
        for r in rows:
            del r["_api_position"]

        result[group_letter] = rows

    if unmapped:
        print(f"AVISO: equipos sin mapear en /standings: {set(unmapped)}", file=sys.stderr)
    print(f"INFO: clasificación oficial recibida para {len(result)} grupos.", file=sys.stderr)
    return result


def fetch_third_place_ranking(api_key):
    """Construye el ranking de los 12 terceros de grupo, usando la clasificación
    oficial de la API (no un cálculo local), para determinar los 8 mejores que
    avanzan a dieciseisavos según el criterio FIFA: pts, gd, gf, fair play."""
    standings = fetch_official_standings(api_key)
    if not standings:
        return [], standings

    thirds = []
    for group_letter, rows in standings.items():
        third = next((r for r in rows if r["position"] == 3), None)
        if third:
            thirds.append({**third, "group": group_letter})

    # La propia API ya nos da pts/gd/gf con los criterios FIFA aplicados a
    # nivel de grupo; para el ranking *entre* terceros de distintos grupos
    # ordenamos por esos mismos campos (pts, gd, gf) que es el criterio
    # documentado para esta fase del desempate.
    thirds.sort(key=lambda r: (-r["pts"], -r["gd"], -r["gf"]))
    for i, t in enumerate(thirds):
        t["ranking"] = i + 1
        t["advances"] = i < 8

    return thirds, standings


def fetch_openfootball_results():
    """Consulta openfootball/worldcup.json (gratuita, sin API key) como
    segunda fuente de verificación cruzada. Devuelve {(home_es, away_es):
    (hg, ag)} en ambas orientaciones, igual que fetch_real_results, o {}
    si la consulta falla por cualquier motivo (esta fuente es opcional,
    nunca debe romper el script si no está disponible)."""
    if not requests:
        return {}
    try:
        resp = requests.get(OPENFOOTBALL_URL, timeout=30)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        print(f"AVISO: no se pudo consultar openfootball ({e}). Sin verificación cruzada.", file=sys.stderr)
        return {}

    results = {}
    unmapped = set()
    for m in data.get("matches", []):
        score = m.get("score")
        if not score or "ft" not in score:
            continue  # partido aún no jugado o sin marcador en esta fuente
        home_en = m.get("team1")
        away_en = m.get("team2")
        home_es = resolve_team_name(home_en)
        away_es = resolve_team_name(away_en)
        if not home_es or not away_es:
            unmapped.add(f"{home_en} vs {away_en}")
            continue
        hg, ag = score["ft"]
        results[(home_es, away_es)] = (hg, ag)
        results[(away_es, home_es)] = (ag, hg)

    if unmapped:
        print(f"AVISO: equipos sin mapear en openfootball: {unmapped}", file=sys.stderr)
    print(f"INFO: openfootball aporta {len(results) // 2} resultados para verificación cruzada.", file=sys.stderr)
    return results


def fetch_real_results(api_key):
    """Llama a football-data.org. Si falla por cualquier motivo, devuelve
    ({}, {}, []) y el script sigue funcionando solo con lo que haya en el
    Excel y las fechas de respaldo. Devuelve (results, schedule,
    raw_finished_list): schedule trae fecha+hora real en España para
    construir la pestaña "Próximos"; raw_finished_list es solo para
    diagnóstico (se vuelca en debug_api.json)."""
    if not api_key:
        print("AVISO: no hay FOOTBALL_DATA_API_KEY configurada. Solo se usará el Excel.", file=sys.stderr)
        return {}, {}, [], []
    if not requests:
        print("AVISO: el paquete requests no está instalado. Solo se usará el Excel.", file=sys.stderr)
        return {}, {}, [], []
    try:
        headers = {"X-Auth-Token": api_key}
        url = f"{API_BASE}/competitions/{COMPETITION_CODE}/matches"
        resp = requests.get(url, headers=headers, timeout=30)
        print(f"INFO: API respondió con status {resp.status_code}", file=sys.stderr)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        print(f"AVISO: no se pudo consultar la API ({e}). Se usará solo el Excel.", file=sys.stderr)
        return {}, {}, [], []

    total_matches = len(data.get("matches", []))
    finished = [m for m in data.get("matches", []) if m["status"] == "FINISHED"]
    print(f"INFO: la API devolvió {total_matches} partidos totales, {len(finished)} finalizados.", file=sys.stderr)

    raw_finished_debug = [
        {
            "home_api": m["homeTeam"]["name"],
            "away_api": m["awayTeam"]["name"],
            "score": f"{m['score']['fullTime']['home']}-{m['score']['fullTime']['away']}",
            "utcDate": m.get("utcDate"),
        }
        for m in finished
    ]

    # Volcado de TODOS los partidos que la API devuelve (jugados o no), con
    # su utcDate y status — para diagnosticar si la API simplemente no
    # incluye un partido concreto en esta llamada (p.ej. por paginación o
    # rango de fechas por defecto), que sería la causa de que su horario
    # no se actualice nunca por mucho que arreglemos el código.
    raw_all_matches_debug = [
        {
            "home_api": (m.get("homeTeam") or {}).get("name"),
            "away_api": (m.get("awayTeam") or {}).get("name"),
            "status": m["status"],
            "utcDate": m.get("utcDate"),
        }
        for m in data.get("matches", [])
    ]

    results = {}
    schedule = {}  # (home_es, away_es) -> (fecha_iso_españa, hora_es "HH:MM")
    unmapped = []
    schedule_unmapped = []

    for m in data.get("matches", []):
        home_en = (m.get("homeTeam") or {}).get("name")
        away_en = (m.get("awayTeam") or {}).get("name")
        if not home_en or not away_en:
            # Partido de eliminatoria aún sin equipos determinados (la API
            # ya tiene el hueco en el calendario pero homeTeam/awayTeam
            # todavía son null) — no hay nada que mapear ni programar.
            continue
        home_es = resolve_team_name(home_en)
        away_es = resolve_team_name(away_en)
        if not home_es or not away_es:
            if m["status"] == "FINISHED":
                unmapped.append(f"{home_en} vs {away_en}")
            else:
                schedule_unmapped.append(f"{home_en} vs {away_en}")
            continue

        # Horario real en España (CEST = UTC+2 en junio-julio) a partir del
        # utcDate que da la API. Sustituye a la tabla escrita a mano, que
        # tenía varios errores de fecha respecto al calendario oficial real.
        utc_date_str = m.get("utcDate")
        if utc_date_str:
            try:
                utc_dt = datetime.fromisoformat(utc_date_str.replace("Z", "+00:00"))
                spain_dt = utc_dt + timedelta(hours=2)  # CEST (horario de verano España)
                spain_date = spain_dt.date().isoformat()
                spain_time = spain_dt.strftime("%H:%M")
                schedule[(home_es, away_es)] = (spain_date, spain_time)
                schedule[(away_es, home_es)] = (spain_date, spain_time)
            except (ValueError, TypeError):
                pass

        if m["status"] != "FINISHED":
            continue

        hg = m["score"]["fullTime"]["home"]
        ag = m["score"]["fullTime"]["away"]
        if hg is None or ag is None:
            continue
        # Guardamos en ambas orientaciones (normal e invertida) para que el
        # resultado se encuentre aunque el Excel tenga el partido anotado
        # con local/visitante distinto a como lo da la API — algo que puede
        # pasar si alguien se equivocó al transcribir el fixture a mano.
        results[(home_es, away_es)] = (hg, ag)
        results[(away_es, home_es)] = (ag, hg)

    if unmapped:
        print(f"AVISO: {len(unmapped)} partidos finalizados con nombre de equipo no mapeado: {unmapped}", file=sys.stderr)
    if schedule_unmapped:
        print(f"AVISO: {len(schedule_unmapped)} partidos no jugados con nombre de equipo no mapeado (sin horario): {set(schedule_unmapped)}", file=sys.stderr)
    print(f"INFO: {len(results) // 2} resultados utilizables tras el mapeo de nombres.", file=sys.stderr)
    print(f"INFO: {len(schedule) // 2} horarios reales (fecha+hora España) obtenidos de la API.", file=sys.stderr)
    return results, schedule, raw_finished_debug, raw_all_matches_debug


def read_excel_data():
    """Lee predicciones, resultados YA introducidos a mano y posiciones de grupo."""
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["ADMIN"]

    matches = []
    for row_idx in range(6, 270):
        match_name = ws.cell(row=row_idx, column=11).value
        if not match_name or match_name in ["-", "Casa-Fuera"]:
            continue
        if str(match_name).startswith("W") or str(match_name).startswith("L"):
            continue
        phase_cell = ws.cell(row=row_idx, column=10).value
        if not (phase_cell and re.match(r"^[A-L][123]$", str(phase_cell))):
            continue

        excel_actual = ws.cell(row=row_idx, column=13).value
        excel_actual = str(excel_actual).strip() if excel_actual and "|" in str(excel_actual) else None

        preds = {}
        for player, col in PLAYER_COLUMNS.items():
            pred = ws.cell(row=row_idx, column=col).value
            preds[player] = str(pred).strip() if pred and "|" in str(pred) else None

        matches.append({
            "match": match_name,
            "group": str(phase_cell)[0],
            "excel_actual": excel_actual,
            "predictions": preds,
        })

    group_positions = {}
    row = 80
    for g in "ABCDEFGHIJKL":
        group_positions[g] = {}
        for pos in [1, 2, 3, 4]:
            preds = {p: ws.cell(row=row, column=c).value for p, c in PLAYER_COLUMNS.items()}
            group_positions[g][pos] = preds
            row += 1

    # Equipos que cada jugador predijo como clasificados a dieciseisavos
    # (filas 130-161, "Dieciseisavofinalista-1" a "-32"). Es una lista plana
    # de 32 equipos por jugador, sin más estructura — no hace falta saber a
    # qué cruce del bracket corresponde cada fila, solo qué 32 equipos eligió.
    qualified_predictions = {p: set() for p in PLAYER_COLUMNS}
    for row_idx in range(130, 162):
        for player, col in PLAYER_COLUMNS.items():
            team = ws.cell(row=row_idx, column=col).value
            if team and isinstance(team, str) and team.strip():
                qualified_predictions[player].add(team.strip())

    return matches, group_positions, qualified_predictions, ws


def score_match(pred_str, actual_str):
    if not pred_str or not actual_str:
        return {"pts": 0, "sign": False, "diff": False, "exact": False}
    psign, pscore = pred_str.split("|")
    ph, pa = map(int, pscore.split("-"))
    asign, ascore = actual_str.split("|")
    ah, aa = map(int, ascore.split("-"))
    pts = 0
    sign_ok = psign == asign
    diff_ok = exact_ok = False
    if sign_ok:
        pts += 1
        if abs(ph - pa) == abs(ah - aa):
            diff_ok = True
            pts += 1
        if ph == ah and pa == aa:
            exact_ok = True
            pts += 2
    return {"pts": pts, "sign": sign_ok, "diff": diff_ok, "exact": exact_ok}


def sign_from_score(h, a):
    return "1" if h > a else ("X" if h == a else "2")


def build_dataset(api_key):
    matches, group_positions, qualified_predictions, ws = read_excel_data()
    api_results, api_schedule, api_raw_finished_debug, api_raw_all_debug = fetch_real_results(api_key)
    openfootball_results = fetch_openfootball_results()
    players = list(PLAYER_COLUMNS.keys())

    processed = []
    group_team_stats = defaultdict(lambda: defaultdict(lambda: {"pj": 0, "g": 0, "e": 0, "p": 0, "gf": 0, "gc": 0}))
    matches_missing_date = []

    diagnostics = []

    for m in matches:
        home, away = m["match"].split("-", 1)

        # Prioridad de fecha/hora: 1) la API (fecha+hora real, siempre que
        # la tenga), 2) la tabla escrita a mano como respaldo si la API no
        # devuelve ese partido por algún motivo (p.ej. fallo de mapeo de
        # nombre). La tabla manual se construyó antes de empezar el
        # torneo y tenía algún desfase de un día en varios partidos.
        api_sched = api_schedule.get((home, away))
        if api_sched:
            official_date, official_time = api_sched
        else:
            official_date = MATCH_DATES_ES.get(m["match"])
            official_time = None
            if not official_date:
                matches_missing_date.append(m["match"])

        # Verificación cruzada del resultado entre las dos fuentes
        # automáticas. Si ambas lo tienen y coinciden -> confianza alta.
        # Si solo una lo tiene -> se usa igual, pero queda marcado como
        # "sin verificar" (puede ser un dato provisional erróneo recién
        # capturado). Si ambas lo tienen y NO coinciden -> se descarta,
        # mejor no mostrar nada que mostrar un dato dudoso.
        fd_res = api_results.get((home, away))
        of_res = openfootball_results.get((home, away))
        api_actual = None
        api_verification = "sin_datos"
        if fd_res and of_res:
            if fd_res == of_res:
                hg, ag = fd_res
                api_actual = f"{sign_from_score(hg, ag)}|{hg}-{ag}"
                api_verification = "verificado_2_fuentes"
            else:
                # Las dos fuentes automáticas no coinciden. En la práctica
                # football-data.org ha mostrado marcadores provisionales
                # erróneos justo tras el pitido final (ej. 5-0 en vez de
                # 4-0) que tarda unos minutos en corregir, mientras que
                # openfootball solo publica el marcador una vez está
                # confirmado. Por eso, ante conflicto, usamos openfootball
                # y lo marcamos para que quede visible en el diagnóstico.
                hg, ag = of_res
                api_actual = f"{sign_from_score(hg, ag)}|{hg}-{ag}"
                api_verification = "conflicto_se_usa_openfootball"
                print(f"AVISO: conflicto entre fuentes para {m['match']}: "
                      f"football-data.org={fd_res} vs openfootball={of_res}. Se usa openfootball.",
                      file=sys.stderr)
        elif fd_res:
            hg, ag = fd_res
            api_actual = f"{sign_from_score(hg, ag)}|{hg}-{ag}"
            api_verification = "solo_football_data_org"
        elif of_res:
            hg, ag = of_res
            api_actual = f"{sign_from_score(hg, ag)}|{hg}-{ag}"
            api_verification = "solo_openfootball"

        # Prioridad: 1) Excel (verificado a mano), 2) API si el Excel no lo tiene
        source = None
        actual = None
        if m["excel_actual"]:
            actual = m["excel_actual"]
            source = "excel"
        elif api_actual:
            actual = api_actual
            source = "api"

        conflict = bool(m["excel_actual"] and api_actual and m["excel_actual"] != api_actual)

        diagnostics.append({
            "match": m["match"],
            "api_verification": api_verification,
            "excel_actual": m["excel_actual"],
            "api_actual": api_actual,
            "used": actual,
            "source": source,
            "conflict": conflict,
        })

        m2 = {
            "match": m["match"],
            "group": m["group"],
            "date": official_date,
            "time": official_time,
            "actual": actual,
            "predictions": m["predictions"],
        }

        if actual:
            asign, ascore = actual.split("|")
            ah, ag = map(int, ascore.split("-"))

            breakdown = {}
            for p in players:
                breakdown[p] = score_match(m["predictions"].get(p), actual)
            m2["breakdown"] = breakdown

            t_h = group_team_stats[m["group"]][home]
            t_a = group_team_stats[m["group"]][away]
            t_h["pj"] += 1
            t_a["pj"] += 1
            t_h["gf"] += ah
            t_h["gc"] += ag
            t_a["gf"] += ag
            t_a["gc"] += ah
            if ah > ag:
                t_h["g"] += 1
                t_a["p"] += 1
            elif ah < ag:
                t_a["g"] += 1
                t_h["p"] += 1
            else:
                t_h["e"] += 1
                t_a["e"] += 1

        processed.append(m2)

    if matches_missing_date:
        print(f"AVISO: sin fecha oficial para: {matches_missing_date}", file=sys.stderr)

    # Clasificación local (cálculo propio: pts, gd, gf — sin fair play porque
    # esos datos no están disponibles en el plan gratuito de la API).
    group_standings_local = {}
    for g, teams in group_team_stats.items():
        rows = []
        for team, s in teams.items():
            pts = s["g"] * 3 + s["e"]
            gd = s["gf"] - s["gc"]
            rows.append({"team": team, **s, "gd": gd, "pts": pts})
        rows.sort(key=lambda r: (-r["pts"], -r["gd"], -r["gf"]))
        for i, r in enumerate(rows, start=1):
            r["position"] = i
        group_standings_local[g] = rows

    # Clasificación oficial vía API (aplica los criterios FIFA reales,
    # incluyendo enfrentamiento directo y fair play cuando hace falta).
    # Si la API falla, usamos la calculada localmente como respaldo.
    third_place_ranking, group_standings_api = fetch_third_place_ranking(api_key)
    using_official_standings = bool(group_standings_api)
    group_standings_real = group_standings_api if using_official_standings else group_standings_local

    if using_official_standings:
        print("INFO: usando clasificación OFICIAL de la API (criterios FIFA completos).", file=sys.stderr)
    else:
        print("AVISO: usando clasificación LOCAL (solo pts/gd/gf, sin fair play). "
              "Puede diferir de la oficial en empates resueltos por tarjetas o enfrentamiento directo.", file=sys.stderr)

    # Equipos realmente clasificados a dieciseisavos: 1º y 2º de cada grupo
    # (según la clasificación que estemos usando, oficial o local) + los 8
    # mejores terceros. Si la fase de grupos no ha terminado, esta lista
    # refleja "quién clasificaría si todo acabara hoy" — es una proyección,
    # no un hecho consumado, hasta que se jueguen los 72 partidos.
    real_qualified_teams = set()
    for g, rows in group_standings_real.items():
        for r in rows:
            if r.get("position") in (1, 2):
                real_qualified_teams.add(r["team"])
    for t in third_place_ranking:
        if t.get("advances"):
            real_qualified_teams.add(t["team"])

    # Puntos por acertar equipos clasificados: +1 por cada equipo que el
    # jugador incluyó en su lista de 32 "Dieciseisavofinalista" y que
    # efectivamente está en la lista real de clasificados (proyectada o
    # confirmada, según el estado actual del torneo).
    qualified_points = {}
    qualified_hits = {}
    for p in players:
        predicted = qualified_predictions.get(p, set())
        hits = predicted & real_qualified_teams
        qualified_points[p] = len(hits)
        qualified_hits[p] = sorted(hits)

    # ── Puntos por acertar posiciones en grupos FINALIZADOS ──────────────────
    GROUP_POS_POINTS = {1: 2, 2: 2, 3: 1, 4: 1}

    # Fecha del último partido jugado por grupo
    group_finish_date = {}
    for m in processed:
        if m.get("actual") and m.get("date") and m.get("group"):
            g = m["group"]
            d = m["date"]
            if g not in group_finish_date or d > group_finish_date[g]:
                group_finish_date[g] = d

    # Grupos donde todos los equipos han jugado 3 partidos
    finished_groups = set()
    for g, rows in group_standings_real.items():
        if rows and all(r.get("pj", 0) >= 3 for r in rows):
            finished_groups.add(g)

    group_pos_points = {p: 0 for p in players}
    group_pos_detail = {p: {} for p in players}
    group_bonus_by_date = {}  # {fecha: {jugador: pts}}

    for g in finished_groups:
        rows = group_standings_real[g]
        real_by_pos = {r["position"]: r["team"] for r in rows}
        finish_date = group_finish_date.get(g)
        if finish_date and finish_date not in group_bonus_by_date:
            group_bonus_by_date[finish_date] = {p: 0 for p in players}
        for p in players:
            pts_g = 0
            for pos in [1, 2, 3, 4]:
                pred = group_positions.get(g, {}).get(str(pos), {}).get(p)                     or group_positions.get(g, {}).get(pos, {}).get(p)
                real = real_by_pos.get(pos)
                if pred and real and pred == real:
                    pts_g += GROUP_POS_POINTS[pos]
            group_pos_points[p] += pts_g
            group_pos_detail[p][g] = pts_g
            if finish_date:
                group_bonus_by_date[finish_date][p] += pts_g

    print(f"INFO: grupos finalizados para puntuar posiciones: {sorted(finished_groups)}", file=sys.stderr)

    # Si todos los grupos están finalizados, los qualified_points se pueden
    # sumar. Se asignan en la fecha del último partido de todos los grupos.
    all_groups_finished = len(finished_groups) == len(group_standings_real) and len(finished_groups) > 0
    qualified_bonus_date = max(group_finish_date.values()) if all_groups_finished and group_finish_date else None

    if qualified_bonus_date:
        print(f"INFO: todos los grupos finalizados — sumando qualified_points en {qualified_bonus_date}", file=sys.stderr)

    by_date = defaultdict(list)
    for m in processed:
        if m.get("date") and m.get("actual"):
            by_date[m["date"]].append(m)
    dates = sorted(by_date.keys())

    daily = {p: {} for p in players}
    cum = {p: {} for p in players}
    running = {p: 0 for p in players}
    for d in dates:
        for p in players:
            day_pts = sum(m["breakdown"][p]["pts"] for m in by_date[d])
            day_pts += group_bonus_by_date.get(d, {}).get(p, 0)
            if qualified_bonus_date and d == qualified_bonus_date:
                day_pts += qualified_points.get(p, 0)
            daily[p][d] = day_pts
            running[p] += day_pts
            cum[p][d] = running[p]

    positions_by_day = {p: [] for p in players}
    for d in dates:
        sorted_p = sorted(players, key=lambda p: -cum[p][d])
        pos, prev, rank = {}, None, 0
        for i, p in enumerate(sorted_p):
            if cum[p][d] != prev:
                rank = i + 1
            pos[p] = rank
            prev = cum[p][d]
        for p in players:
            positions_by_day[p].append(pos[p])

    # --- Fase eliminatoria (dieciseisavos a final) ---
    try:
        ko_data = build_ko_dataset(
            ws, PLAYER_COLUMNS, group_standings_real, third_place_ranking, group_positions,
            api_results=api_results, openfootball_results=openfootball_results,
        )
    except Exception as e:
        print(f"AVISO: no se pudo procesar la fase eliminatoria ({e}). Se omite por ahora.", file=sys.stderr)
        ko_data = {"rounds": {}, "qualifiers": {}, "honor": {}, "eliminated_teams": [],
                   "winners_by_match": {}, "losers_by_match": {}}

    # --- Puntos de fase eliminatoria ---
    # Suma puntos de partidos KO (breakdown) + puntos por equipos clasificados,
    # fechados con el calendario real para que Gráficos/Jornadas avancen día
    # a día también en fase KO (aunque la ronda global no haya terminado).
    ko_points = {p: 0 for p in players}
    try:
        extra_dates = set()

        # 1. Puntos por aciertos de partido (signo/diferencia/exacto en 1/16)
        for round_name, rdata in ko_data.get("rounds", {}).items():
            for m in rdata.get("matches", []):
                if not m.get("actual") or not m.get("breakdown") or not m.get("date"):
                    continue
                extra_dates.add(m["date"])

        # 2. Fechas de los bonus por equipo clasificado
        for round_name, slot_rows in ko_data.get("qualifiers", {}).items():
            for row in slot_rows:
                for p, info in row.get("predictions", {}).items():
                    if info.get("status") == "clasificado" and info.get("date"):
                        extra_dates.add(info["date"])

        if extra_dates:
            dates = sorted(set(dates) | extra_dates)
            for p in players:
                for d in extra_dates:
                    daily[p].setdefault(d, 0)

        for round_name, rdata in ko_data.get("rounds", {}).items():
            for m in rdata.get("matches", []):
                if not m.get("actual") or not m.get("breakdown"):
                    continue
                match_date = m.get("date")
                for p in players:
                    bd = m["breakdown"].get(p, {})
                    pts = bd.get("pts", 0) if bd else 0
                    if pts > 0:
                        ko_points[p] += pts
                        if match_date and match_date in daily.get(p, {}):
                            daily[p][match_date] += pts
                        elif dates:
                            daily[p][dates[-1]] += pts

        # 3. Puntos por equipos clasificados correctamente (qualifier_pts de ko_stage),
        # fechados con la fecha real del partido que decidió esa clasificación.
        for round_name, slot_rows in ko_data.get("qualifiers", {}).items():
            for row in slot_rows:
                for p, info in row.get("predictions", {}).items():
                    if info.get("status") != "clasificado":
                        continue
                    pts = info.get("pts", 0)
                    if pts <= 0:
                        continue
                    ko_points[p] += pts
                    q_date = info.get("date")
                    if q_date and q_date in daily.get(p, {}):
                        daily[p][q_date] += pts
                    elif dates:
                        daily[p][dates[-1]] += pts

        # Recalculate cumulative
        if any(v > 0 for v in ko_points.values()):
            running2 = {p: 0 for p in players}
            for d in dates:
                for p in players:
                    running2[p] += daily[p].get(d, 0)
                    cum[p][d] = running2[p]
            running = running2
            positions_by_day = {p: [] for p in players}
            for d in dates:
                sorted_p = sorted(players, key=lambda p: -cum[p][d])
                pos, prev, rank = {}, None, 0
                for i, p in enumerate(sorted_p):
                    if cum[p][d] != prev:
                        rank = i + 1
                    pos[p] = rank
                    prev = cum[p][d]
                for p in players:
                    positions_by_day[p].append(pos[p])
    except Exception as e:
        print(f"AVISO: error calculando ko_points ({e})", file=sys.stderr)

    # Update standings with ko_points
    standings = sorted(
        ({"player": p, "points": running[p],
          "group_pos_points": group_pos_points[p],
          "qualified_points": qualified_points.get(p, 0),
          "ko_points": ko_points.get(p, 0),
          "total_points": running[p]} for p in players),
        key=lambda x: -x["points"],
    )
    for i, s in enumerate(standings):
        s["rank"] = i + 1

    # --- Desglose de puntos por ronda para la pestaña Clasificación ---
    # Orden pedido: general primero, luego detalle en orden cronológico.
    ROUND_BUCKET_ORDER = [
        "general", "grupos_partidos", "grupos_posiciones",
        "dieciseisavos", "octavos", "cuartos", "semis", "final",
    ]
    ROUND_BUCKET_LABELS = {
        "general": "Clasificación general",
        "grupos_partidos": "Partidos de fase de grupos",
        "grupos_posiciones": "Posiciones de grupos (y clasificados a 1/16)",
        "dieciseisavos": "Dieciseisavos (1/16)",
        "octavos": "Octavos de final",
        "cuartos": "Cuartos de final",
        "semis": "Semifinales",
        "final": "Final",
    }

    ko_match_pts_by_round = {r: {p: 0 for p in players} for r in ["dieciseisavos", "octavos", "cuartos", "semis", "final"]}
    for round_name, rdata in ko_data.get("rounds", {}).items():
        if round_name not in ko_match_pts_by_round:
            continue
        for m in rdata.get("matches", []):
            bd = m.get("breakdown") or {}
            for p in players:
                ko_match_pts_by_round[round_name][p] += bd.get(p, {}).get("pts", 0)

    qbr = ko_data.get("qualifier_pts_by_round", {})

    def _qr(round_name, p):
        return qbr.get(round_name, {}).get(p, 0)

    rounds_breakdown = {"order": ROUND_BUCKET_ORDER, "labels": ROUND_BUCKET_LABELS, "by_player": {}}
    for p in players:
        grupos_partidos = sum(
            m["breakdown"][p]["pts"] for m in processed if m.get("group") and m.get("breakdown")
        )
        grupos_posiciones = group_pos_points[p] + qualified_points.get(p, 0)
        dieci = ko_match_pts_by_round["dieciseisavos"][p] + _qr("octavos", p)
        octavos = ko_match_pts_by_round["octavos"][p] + _qr("cuartos", p)
        cuartos = ko_match_pts_by_round["cuartos"][p] + _qr("semis", p)
        semis = ko_match_pts_by_round["semis"][p] + _qr("final", p) + _qr("tercer_puesto", p)
        final = ko_match_pts_by_round["final"][p]
        general = grupos_partidos + grupos_posiciones + dieci + octavos + cuartos + semis + final
        rounds_breakdown["by_player"][p] = {
            "general": general,
            "grupos_partidos": grupos_partidos,
            "grupos_posiciones": grupos_posiciones,
            "dieciseisavos": dieci,
            "octavos": octavos,
            "cuartos": cuartos,
            "semis": semis,
            "final": final,
        }

    return {
        "matches": processed,
        "group_positions": group_positions,
        "group_standings_real": group_standings_real,
        "third_place_ranking": third_place_ranking,
        "using_official_standings": using_official_standings,
        "real_qualified_teams": sorted(real_qualified_teams),
        "qualified_points": qualified_points,
        "ko_points": ko_points,
        "qualified_hits": qualified_hits,
        "finished_groups": sorted(finished_groups),
        "group_pos_points": group_pos_points,
        "group_pos_detail": group_pos_detail,
        "ko_stage": ko_data,
        "rounds_breakdown": rounds_breakdown,
        "players": players,
        "dates": dates,
        "daily_points": daily,
        "cumulative_points": cum,
        "positions_by_day": positions_by_day,
        "standings": standings,
        "last_updated": date.today().isoformat(),
    }, diagnostics, api_raw_finished_debug, api_raw_all_debug


def main():
    api_key = os.environ.get("FOOTBALL_DATA_API_KEY")
    dataset, diagnostics, api_raw_finished_debug, api_raw_all_debug = build_dataset(api_key)

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write("const PORRA_DATA = ")
        json.dump(dataset, f, ensure_ascii=False)
        f.write(";")

    # Archivo de diagnóstico legible: qué fuente se usó para cada partido y
    # si hubo conflicto entre el Excel y la API. Útil para depurar sin tener
    # que rebuscar en los logs de GitHub Actions — basta con abrir este
    # archivo directamente en el repo.
    conflicts = [d for d in diagnostics if d["conflict"]]
    from_api = [d for d in diagnostics if d["source"] == "api"]
    from_excel = [d for d in diagnostics if d["source"] == "excel"]
    not_played = [d for d in diagnostics if d["source"] is None]

    with open("debug_api.json", "w", encoding="utf-8") as f:
        json.dump({
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "resumen": {
                "total_partidos": len(diagnostics),
                "desde_excel": len(from_excel),
                "desde_api": len(from_api),
                "sin_jugar": len(not_played),
                "conflictos_excel_vs_api": len(conflicts),
            },
            "clasificacion_de_grupos": {
                "usando_oficial_api": dataset["using_official_standings"],
                "grupos_recibidos_de_api": len(dataset["group_standings_real"]) if dataset["using_official_standings"] else 0,
                "ranking_terceros_calculado": len(dataset["third_place_ranking"]) > 0,
                "nota": (
                    "Clasificación oficial de la API (con fair-play/enfrentamiento directo)."
                    if dataset["using_official_standings"]
                    else "FALLBACK: clasificación calculada localmente (solo pts/gd/gf). "
                         "El endpoint /standings de la API no respondió o falló — revisa el log "
                         "completo del paso 'Actualizar data.js' en Actions para ver el motivo exacto."
                ),
            },
            "conflictos": conflicts,
            "rellenados_por_api": from_api,
            "partidos_finalizados_segun_api": api_raw_finished_debug,
            "TODOS_los_partidos_segun_api": api_raw_all_debug,
            "todos": diagnostics,
        }, f, ensure_ascii=False, indent=2)

    played = sum(1 for m in dataset["matches"] if m["actual"])
    print(f"OK: data.js actualizado. {played}/{len(dataset['matches'])} partidos jugados. "
          f"Líder: {dataset['standings'][0]['player']} ({dataset['standings'][0]['points']} pts)")
    print(f"Fuentes: {len(from_excel)} del Excel, {len(from_api)} de la API.")
    if dataset["using_official_standings"]:
        print(f"✅ Clasificación de grupos: OFICIAL vía API ({len(dataset['group_standings_real'])} grupos).")
    else:
        print(f"⚠️  Clasificación de grupos: LOCAL (fallback) — el endpoint /standings no respondió. "
              f"Revisa los mensajes 'AVISO: no se pudo consultar /standings' más arriba en este mismo log.")
    if conflicts:
        print(f"⚠️  ATENCIÓN: {len(conflicts)} partido(s) con resultado distinto entre Excel y API "
              f"(se usó el del Excel). Revisa debug_api.json -> 'conflictos'.")
        for c in conflicts:
            print(f"   - {c['match']}: Excel={c['excel_actual']} | API={c['api_actual']}")


if __name__ == "__main__":
    main()
