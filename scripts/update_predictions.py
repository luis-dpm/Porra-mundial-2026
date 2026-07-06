#!/usr/bin/env python3
"""Genera predictions_data.js: proyección probabilística de lo que queda de
torneo (desde los octavos sin jugar hasta la final), combinando:

  - El estado ya calculado por update_data.py en data.js (standings,
    ko_stage.rounds / qualifiers / eliminated_teams / winners_by_match).
  - Dos columnas del Excel de reglas que update_data.py todavía no vuelca a
    JSON: "Subcampeón" (fila 251) y "3º puesto" ganador (fila 252). Se leen
    aquí en modo solo-lectura; este script no toca data.js ni el resto del
    pipeline de resultados reales.
  - Datos de mercado externos que no existen en el Excel: cuotas "to advance"
    de Kalshi para los octavos sin jugar y para Cuartos 1 (Marruecos-Francia,
    ya es un cruce real), rating Elo de cada selección y probabilidades de
    Polymarket para Bota de Oro / Balón de Oro. Consultados 4-5 jul 2026.

No se ejecuta en el workflow automático (update.yml) porque depende de estos
datos de mercado, que hay que refrescar a mano cuando cambian. Antes de cada
refresco, corre `python3 scripts/check_market_data.py` para saber exactamente
qué cuotas faltan o ya se pueden borrar según cómo va el torneo en data.js.
"""
import json
import re
import sys
import itertools
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")
import openpyxl

ROOT = Path(__file__).resolve().parent.parent

PLAYER_COLUMNS = {
    "LUIS DPM": 19, "EL MATO": 22, "IVÁN DELGADO": 25, "ADRIÁN": 28,
    "JUAN": 31, "CARLOS": 34, "SU FLORENTINEZA": 37,
}

# ---------------------------------------------------------------- mercado --
# Octavos sin jugar: cuota real "to advance" de Kalshi (consultada 6 jul. 2026).
# Clave = número de partido FIFA.
OCTAVOS_ODDS = {
    93: {"España": 0.66, "Portugal": 0.34},
    94: {"Bélgica": 0.47, "Estados Unidos": 0.53},
    95: {"Argentina": 0.84, "Egipto": 0.16},
    96: {"Colombia": 0.60, "Suiza": 0.40},
}
# Cruces ya confirmados (los dos equipos conocidos) con mercado propio de
# Kalshi, más allá de los octavos -- se va ampliando ronda a ronda según se
# van jugando los octavos/cuartos/semis reales. Clave = frozenset con los dos
# equipos; valor = precio "to advance" de cada uno (no hace falta que sumen 1,
# hybrid_prob() ya normaliza). Ver scripts/check_market_data.py para saber qué
# cruce hay que añadir aquí en cada momento.
KNOWN_MATCHUPS = {
    frozenset({"Francia", "Marruecos"}): {"Francia": 0.76, "Marruecos": 0.24},  # Cuartos 1, partido 97, consultada 6 jul. 2026
    frozenset({"Noruega", "Inglaterra"}): {"Noruega": 0.36, "Inglaterra": 0.64},  # Cuartos 3, partido 99, consultada 6 jul. 2026
}

# World Football Elo (eloratings.net/2026_World_Cup), ratings al lunes 6 jul.
# 2026. Ancla estable para cruces que todavía no existen como mercado (cuartos
# 2-4, semis, final, 3º-4º puesto).
ELO = {
    "España": 2159, "Argentina": 2151, "Francia": 2143, "Inglaterra": 2046,
    "Brasil": 1993, "Portugal": 2013, "Colombia": 2009, "México": 1943,
    "Suiza": 1943, "Noruega": 1972, "Marruecos": 1921, "Bélgica": 1910,
    "Paraguay": 1814, "Estados Unidos": 1798, "Canadá": 1729, "Egipto": 1747,
}

def elo_prob(a, b):
    return 1.0 / (1.0 + 10 ** (-(ELO[a] - ELO[b]) / 400.0))

def hybrid_prob(a, b):
    """Cuota real de mercado si el cruce ya existe como partido concreto (ver
    KNOWN_MATCHUPS), si no cae al Elo."""
    key = frozenset({a, b})
    if key in KNOWN_MATCHUPS:
        odds = KNOWN_MATCHUPS[key]
        return odds[a] / (odds[a] + odds[b])
    return elo_prob(a, b)

# Bota de Oro (consultada 6 jul. 2026), cifras tal cual las da el mercado
# (sin renormalizar): suman 100.1%, así que "Otros" se queda en 0 -- no hay
# margen que repartirle. Ojo: la clave debe ser "Julián Álvarez" completo,
# que es como aparece el pick en la hoja (JUAN lo tiene picado).
GOLDEN_CANDIDATES = [
    ("Mbappé", 0.49), ("Messi", 0.325), ("Haaland", 0.104),
    ("Kane", 0.068), ("Oyarzabal", 0.013), ("Julián Álvarez", 0.001),
    ("Otros", 0.0),
]
# Balón de Oro (consultada 6 jul. 2026), cifras tal cual (suman 102.1%,
# "Otros" se queda en 0 igual que arriba). Ojo: la clave debe ser "Lamine
# Yamal" completo (no solo "Yamal") y "Declan Rice" completo, que es como
# aparecen esos picks en la hoja (Yamal lo tienen picado 5 de 7 jugadores;
# Declan Rice, SU FLORENTINEZA).
GOLDEN_BALL_CANDIDATES = [
    ("Mbappé", 0.38), ("Messi", 0.26), ("Haaland", 0.093),
    ("Olise", 0.086), ("Kane", 0.068), ("Bellingham", 0.062),
    ("Lamine Yamal", 0.027), ("Ronaldo", 0.021), ("Dembélé", 0.018),
    ("Pedri", 0.005), ("Declan Rice", 0.001), ("Otros", 0.0),
]

# ------------------------------------------------------------- data.js I/O --
def load_porra_data():
    src = (ROOT / "data.js").read_text(encoding="utf-8")
    m = re.search(r"^const PORRA_DATA = (\{.*\});\s*$", src, re.S)
    if not m:
        sys.exit("No se pudo encontrar PORRA_DATA en data.js")
    return json.loads(m.group(1))

def load_award_picks():
    wb = openpyxl.load_workbook(ROOT / "source" / "predicciones.xlsx", data_only=True)
    ws = wb["ADMIN"]
    runner_up = {p: ws.cell(row=251, column=c).value for p, c in PLAYER_COLUMNS.items()}
    third_place_winner = {p: ws.cell(row=252, column=c).value for p, c in PLAYER_COLUMNS.items()}
    return runner_up, third_place_winner

# ------------------------------------------------------- bracket topology --
def ref_num(ref):
    return int(ref[1:])

def build_topology(porra):
    ko = porra["ko_stage"]["rounds"]
    octavos_matches = ko["octavos"]["matches"]
    octavos_num_to_idx = {m["num"]: i for i, m in enumerate(octavos_matches)}
    cuartos_matches = ko["cuartos"]["matches"]
    cuartos_num_to_idx = {m["num"]: i for i, m in enumerate(cuartos_matches)}
    semis_matches = ko["semis"]["matches"]
    semis_num_to_idx = {m["num"]: i for i, m in enumerate(semis_matches)}

    qf_pairs = [
        (octavos_num_to_idx[ref_num(m["home_ref"])], octavos_num_to_idx[ref_num(m["away_ref"])])
        for m in cuartos_matches
    ]
    sf_pairs = [
        (cuartos_num_to_idx[ref_num(m["home_ref"])], cuartos_num_to_idx[ref_num(m["away_ref"])])
        for m in semis_matches
    ]
    final_match = ko["final"]["matches"][0]
    f_pair = (semis_num_to_idx[ref_num(final_match["home_ref"])], semis_num_to_idx[ref_num(final_match["away_ref"])])
    # 3o-4o puesto lo alimentan los MISMOS partidos de semis, pero sus perdedores
    tp_match = ko["tercer_puesto"]["matches"][0]
    tp_pair = (semis_num_to_idx[ref_num(tp_match["home_ref"])], semis_num_to_idx[ref_num(tp_match["away_ref"])])
    return octavos_matches, cuartos_matches, semis_matches, qf_pairs, sf_pairs, f_pair, tp_pair

# ------------------------------------------------------------ octavos data --
# Partidos ya decididos en la realidad pero que data.js todavía no refleja
# (a la espera de que el pipeline automático del sitio recoja el resultado
# oficial con marcador). Clave = número de partido FIFA, valor = ganador.
# Quitar la entrada de aquí en cuanto data.js incorpore el resultado real.
MANUAL_RESULTS = {}

def build_octavos(porra, octavos_matches):
    winners = porra["ko_stage"]["winners_by_match"]
    out = []
    for m in octavos_matches:
        home, away = m["home_team"], m["away_team"]
        if m["actual"]:
            winner = winners[str(m["num"])]
            score = m["actual"].split("|")[1]
            out.append({"num": m["num"], "a": home, "b": away, "resolved": True,
                        "winner": winner, "score": score, "date": m["date"]})
        elif m["num"] in MANUAL_RESULTS:
            out.append({"num": m["num"], "a": home, "b": away, "resolved": True,
                        "winner": MANUAL_RESULTS[m["num"]], "score": "—", "date": m["date"]})
        else:
            odds = OCTAVOS_ODDS[m["num"]]
            out.append({"num": m["num"], "a": home, "b": away, "resolved": False,
                        "probA": odds[home], "probB": odds[away], "date": m["date"]})
    return out

# ------------------------------------------------------------ player picks --
# IMPORTANTE: qualifiers.cuartos/semis en data.js conservan el orden CRUDO de
# filas del Excel ("Cuartofinalista-1..8", "Semifinalista-1..4"), que no
# coincide con el orden de partido real (89..96 / 97..100) que usa esta
# simulación. No hay una tabla fija que traduzca uno a otro (es simplemente
# el orden en que el autor de la hoja puso las filas), así que se deduce por
# mayoría: para cada slot crudo, qué partido real contiene más picks válidos
# de los 7 jugadores (los picks de equipos ya eliminados en fases previas no
# encajan en ningún partido y no cuentan).
def derive_slot_permutation(slots, candidate_team_sets):
    n = len(slots)
    used = set()
    perm = [None] * n
    for i, slot in enumerate(slots):
        picks = [info["team"] for info in slot["predictions"].values() if info["team"]]
        best_j, best_score = None, -1
        for j, teams in enumerate(candidate_team_sets):
            if j in used:
                continue
            score = sum(1 for t in picks if t in teams)
            if score > best_score:
                best_score = score
                best_j = j
        perm[i] = best_j
        used.add(best_j)
    if None in perm or len(set(perm)) != n:
        raise ValueError(f"No se pudo deducir una permutación válida de slots: {perm}")
    return perm

def build_player_picks(porra, runner_up, third_place_winner, octavos, qf_pairs):
    q = porra["ko_stage"]["qualifiers"]
    fp = porra["final_predictions"]

    cuartos_candidates = [{o["a"], o["b"]} for o in octavos]
    cuarto_perm = derive_slot_permutation(q["cuartos"], cuartos_candidates)

    semis_candidates = []
    for ia, ib in qf_pairs:
        oa, ob = octavos[ia], octavos[ib]
        semis_candidates.append({oa["a"], oa["b"], ob["a"], ob["b"]})
    semi_perm = derive_slot_permutation(q["semis"], semis_candidates)

    picks = {}
    for p in porra["players"]:
        raw_c = [slot["predictions"][p]["team"] for slot in q["cuartos"]]
        raw_s = [slot["predictions"][p]["team"] for slot in q["semis"]]
        picks[p] = dict(
            cuartofinalista=[raw_c[cuarto_perm.index(i)] for i in range(8)],
            semifinalista=[raw_s[semi_perm.index(k)] for k in range(4)],
            finalista=[slot["predictions"][p]["team"] for slot in q["final"]],
            tresycuatro=[slot["predictions"][p]["team"] for slot in q["tercer_puesto"]],
            campeon=fp[p]["campeon"], subcampeon=runner_up[p], tercerpuesto=third_place_winner[p],
            botaoro=fp[p]["bota_oro"], balonoro=fp[p]["balon_oro"],
        )
    return picks

def build_dead_list(porra, runner_up, third_place_winner, octavos, cuartos_matches, semis_matches, picks):
    """Un pick solo cuenta como "muerto" si el partido real que decide ESE
    slot concreto todavía no se ha jugado -- es decir, si el puesto sigue en
    juego para el resto de jugadores pero este pick ya es matemáticamente
    imposible. Si el partido que ocupa ese slot ya se jugó (haya acertado el
    jugador o no), ya no es un pick "muerto": es un fallo normal, resuelto,
    y no tiene sentido seguir listándolo (p. ej. un cuartofinalista picado
    que perdió su propio octavos ya jugado -- ese cruce ya tiene ganador real
    y punto). En cambio, un semifinalista ya eliminado en octavos SÍ sigue
    siendo un pick muerto mientras el cruce de cuartos que alimenta ese
    puesto no se haya jugado, porque el puesto en sí sigue abierto para
    otros jugadores."""
    eliminated = set(porra["ko_stage"]["eliminated_teams"])
    dead = {p: [] for p in porra["players"]}

    for p in porra["players"]:
        for i, o in enumerate(octavos):
            team = picks[p]["cuartofinalista"][i]
            if not o["resolved"] and team in eliminated:
                dead[p].append({"stage": "Cuartofinalista", "pick": team, "match": None})
        for k, m in enumerate(cuartos_matches):
            team = picks[p]["semifinalista"][k]
            if not m["actual"] and team in eliminated:
                dead[p].append({"stage": "Semifinalista", "pick": team, "match": None})

    # Finalista, 3º-4º puesto, Campeón, Subcampeón y 3º puesto (ganador) los
    # decide un cruce de semis o la final, que a día de hoy no se ha jugado
    # ninguno todavía -- así que, por ahora, cualquier equipo ya eliminado
    # sigue siendo un pick muerto para estas categorías sin excepción.
    q = porra["ko_stage"]["qualifiers"]
    labels = {"final": "Finalista", "tercer_puesto": "3º-4º puesto (clasificado)"}
    for round_name, label in labels.items():
        for slot in q[round_name]:
            for p, info in slot["predictions"].items():
                if info["status"] == "eliminado":
                    dead[p].append({"stage": label, "pick": info["team"], "match": None})
    for p in porra["players"]:
        if third_place_winner[p] in eliminated:
            dead[p].append({"stage": "3º puesto (ganador)", "pick": third_place_winner[p], "match": None})
        if runner_up[p] in eliminated:
            dead[p].append({"stage": "Subcampeón", "pick": runner_up[p], "match": None})
        if porra["final_predictions"][p]["campeon"] in eliminated:
            dead[p].append({"stage": "Campeón", "pick": porra["final_predictions"][p]["campeon"], "match": None})
    return dead

# --------------------------------------------------------------- simulate --
def make_simulator(octavos, qf_pairs, sf_pairs, f_pair, tp_pair):
    resolved_idx = [i for i, o in enumerate(octavos) if o["resolved"]]
    unresolved_idx = [i for i, o in enumerate(octavos) if not o["resolved"]]
    n_unresolved = len(unresolved_idx)
    n_bits = n_unresolved + len(qf_pairs) + len(sf_pairs) + 1 + 1

    def simulate(bits, mode):
        nu = n_unresolved
        o_bits = bits[0:nu]
        q_bits = bits[nu:nu + 4]
        s_bits = bits[nu + 4:nu + 6]
        f_bit = bits[nu + 6]
        t_bit = bits[nu + 7]

        O_winner = [None] * len(octavos)
        prob = 1.0
        for i in resolved_idx:
            O_winner[i] = octavos[i]["winner"]
        for j, i in enumerate(unresolved_idx):
            o = octavos[i]
            pA = o["probA"] if mode == "weighted" else 0.5
            if o_bits[j] == 0:
                O_winner[i] = o["a"]; prob *= pA
            else:
                O_winner[i] = o["b"]; prob *= (1 - pA)

        Q_winner = []
        for k, (ia, ib) in enumerate(qf_pairs):
            teamA, teamB = O_winner[ia], O_winner[ib]
            pA = hybrid_prob(teamA, teamB) if mode == "weighted" else 0.5
            if q_bits[k] == 0:
                Q_winner.append(teamA); prob *= pA
            else:
                Q_winner.append(teamB); prob *= (1 - pA)

        S_winner, S_loser = [], []
        for k, (ia, ib) in enumerate(sf_pairs):
            teamA, teamB = Q_winner[ia], Q_winner[ib]
            pA = hybrid_prob(teamA, teamB) if mode == "weighted" else 0.5
            if s_bits[k] == 0:
                S_winner.append(teamA); S_loser.append(teamB); prob *= pA
            else:
                S_winner.append(teamB); S_loser.append(teamA); prob *= (1 - pA)

        ia, ib = f_pair
        teamA, teamB = S_winner[ia], S_winner[ib]
        pA = hybrid_prob(teamA, teamB) if mode == "weighted" else 0.5
        if f_bit == 0:
            champion, runner = teamA, teamB; prob *= pA
        else:
            champion, runner = teamB, teamA; prob *= (1 - pA)

        ia, ib = tp_pair
        teamA34, teamB34 = S_loser[ia], S_loser[ib]
        pA34 = hybrid_prob(teamA34, teamB34) if mode == "weighted" else 0.5
        if t_bit == 0:
            winner34 = teamA34; prob *= pA34
        else:
            winner34 = teamB34; prob *= (1 - pA34)

        return O_winner, Q_winner, S_winner, S_loser, champion, runner, winner34, prob

    return simulate, n_bits, unresolved_idx

def score_player_bracket(pk, O_winner, Q_winner, S_winner, S_loser, champion, runner, winner34, scoreable_octavos):
    """scoreable_octavos: índices de octavos que hay que puntuar aquí. Los ya
    resueltos (Francia, Marruecos) se excluyen porque sus 12 pts por acertar
    cuartofinalista ya están sumados en current_total (viene de data.js, que
    ya los contabiliza en cuanto el equipo queda "clasificado") -- puntuarlos
    otra vez aquí los contaría dos veces."""
    s = 0
    for slot in scoreable_octavos:
        if pk["cuartofinalista"][slot] == O_winner[slot]: s += 12
    for slot in range(4):
        if pk["semifinalista"][slot] == Q_winner[slot]: s += 24
    for slot in range(2):
        if pk["finalista"][slot] == S_winner[slot]: s += 48
    for slot in range(2):
        if pk["tresycuatro"][slot] == S_loser[slot]: s += 24
    if pk["campeon"] == champion: s += 96
    if pk["subcampeon"] == runner: s += 48
    if pk["tercerpuesto"] == winner34: s += 16
    return s

PAYMENT_STEPS = [0, 1/3, 2/3, 1, 4/3, 5/3, 2]

def ranks_with_ties(scores):
    order = sorted(scores.items(), key=lambda kv: -kv[1])
    n = len(order); ranks = {}; i = 0
    while i < n:
        j = i
        while j + 1 < n and order[j + 1][1] == order[i][1]: j += 1
        avg_rank = sum(range(i + 1, j + 2)) / (j - i + 1)
        for k in range(i, j + 1): ranks[order[k][0]] = avg_rank
        i = j + 1
    return ranks

def payment_for_rank(avg_rank):
    return (avg_rank - 1) / 3

def rank_groups(scores):
    """Grupos de empate (listas de jugadores), de mayor a menor puntuación."""
    order = sorted(scores.items(), key=lambda kv: -kv[1])
    n = len(order); groups = []; i = 0
    while i < n:
        j = i
        while j + 1 < n and order[j + 1][1] == order[i][1]: j += 1
        groups.append([order[k][0] for k in range(i, j + 1)])
        i = j + 1
    return groups

def wq(pairs, q, tot):
    items = sorted(pairs.items())
    target = q * tot; cum = 0.0
    for val, pr in items:
        cum += pr
        if cum >= target - 1e-9: return val
    return items[-1][0]

# -------------------------------------------------------------------- main --
def main():
    porra = load_porra_data()
    players = porra["players"]
    current_total = {s["player"]: s["points"] for s in porra["standings"]}
    runner_up, third_place_winner = load_award_picks()

    octavos_matches, cuartos_matches, semis_matches, qf_pairs, sf_pairs, f_pair, tp_pair = build_topology(porra)
    octavos = build_octavos(porra, octavos_matches)
    picks = build_player_picks(porra, runner_up, third_place_winner, octavos, qf_pairs)
    dead = build_dead_list(porra, runner_up, third_place_winner, octavos, cuartos_matches, semis_matches, picks)

    match_labels = []
    for o in octavos:
        if not o["resolved"]:
            match_labels.append(("Octavos", o["a"], o["b"]))
    stage_names = {0: "Cuartos", 1: "Cuartos", 2: "Cuartos", 3: "Cuartos"}
    for k in range(4):
        ia, ib = qf_pairs[k]
        a = octavos[ia]["winner"] if octavos[ia]["resolved"] else f"Ganador({octavos[ia]['a']}/{octavos[ia]['b']})"
        b = octavos[ib]["winner"] if octavos[ib]["resolved"] else f"Ganador({octavos[ib]['a']}/{octavos[ib]['b']})"
        match_labels.append(("Cuartos", a, b))
    for k in range(2):
        ia, ib = sf_pairs[k]
        match_labels.append(("Semis", f"Ganador(Cuartos {ia+1})", f"Ganador(Cuartos {ib+1})"))
    match_labels.append(("Final", f"Ganador(Semis {f_pair[0]+1})", f"Ganador(Semis {f_pair[1]+1})"))
    match_labels.append(("3º-4º puesto", f"Perdedor(Semis {tp_pair[0]+1})", f"Perdedor(Semis {tp_pair[1]+1})"))

    simulate, n_bits, unresolved_idx = make_simulator(octavos, qf_pairs, sf_pairs, f_pair, tp_pair)
    all_bits = list(itertools.product([0, 1], repeat=n_bits))
    print(f"bracket branches: {len(all_bits)} ({n_bits} bits)", file=sys.stderr)

    golden_bonus = {}
    for p in players:
        prob_real = dict(GOLDEN_CANDIDATES).get(picks[p]["botaoro"], 0.0)
        ball_prob_real = dict(GOLDEN_BALL_CANDIDATES).get(picks[p]["balonoro"], 0.0)
        golden_bonus[p] = 24 * prob_real + 24 * ball_prob_real

    summary = {}
    chuleton_exp = {}
    chuleton_dist = {}
    matches_out = {}
    bracket_min = {p: None for p in players}
    bracket_max = {p: None for p in players}
    rank_dist_pct = {}
    total_prob_by_mode = {}
    # Mejor escenario (mayor probabilidad conjunta) en el que cada jugador
    # queda 1º (o empatado a 1º), solo en modo ponderado -- en equiprobable
    # todas las ramas del cuadro pesan igual, así que "la más probable" no
    # sería más que una cualquiera de las muchas empatadas al máximo.
    camino_best = {p: {"w": -1.0, "bits": None, "gname": None, "bname": None, "score": None,
                       "tied_with": [], "second_name": None, "second_score": None} for p in players}

    for mode in ["uniform", "weighted"]:
        win_mass = {p: 0.0 for p in players}
        score_sum = {p: 0.0 for p in players}
        score_pairs = {p: {} for p in players}
        cuts_sum = {p: 0.0 for p in players}
        cuts_pairs = {p: {} for p in players}
        n_matches = len(match_labels)
        impact_mass_a = [0.0] * n_matches
        impact_mass_b = [0.0] * n_matches
        impact_cuts_a = [{p: 0.0 for p in players} for _ in range(n_matches)]
        impact_cuts_b = [{p: 0.0 for p in players} for _ in range(n_matches)]
        rank_dist = {p: [0.0] * len(players) for p in players}
        total_prob = 0.0

        for bits in all_bits:
            O_winner, Q_winner, S_winner, S_loser, champion, runner, winner34, bprob = simulate(bits, mode)
            bracket_scores = {p: score_player_bracket(picks[p], O_winner, Q_winner, S_winner, S_loser, champion, runner, winner34, unresolved_idx) for p in players}
            if mode == "uniform":
                for p in players:
                    bracket_min[p] = bracket_scores[p] if bracket_min[p] is None else min(bracket_min[p], bracket_scores[p])
                    bracket_max[p] = bracket_scores[p] if bracket_max[p] is None else max(bracket_max[p], bracket_scores[p])

            for gname, gprob in GOLDEN_CANDIDATES:
                for bname, bprob2 in GOLDEN_BALL_CANDIDATES:
                    w = bprob * gprob * bprob2 if mode == "weighted" else (1.0 / len(all_bits)) * gprob * bprob2
                    total_prob += w
                    scores = {}
                    for p in players:
                        s = current_total[p] + bracket_scores[p]
                        if picks[p]["botaoro"] == gname: s += 24
                        if picks[p]["balonoro"] == bname: s += 24
                        scores[p] = s
                    ranks = ranks_with_ties(scores)
                    maxscore = max(scores.values())
                    leaders = [p for p in players if scores[p] == maxscore]
                    for p in leaders: win_mass[p] += w / len(leaders)
                    if mode == "weighted":
                        groups_for_camino = rank_groups(scores)
                        second_group = groups_for_camino[1] if len(groups_for_camino) > 1 else []
                        for p in leaders:
                            if w > camino_best[p]["w"]:
                                camino_best[p] = {
                                    "w": w, "bits": bits, "gname": gname, "bname": bname, "score": scores[p],
                                    "tied_with": [q for q in leaders if q != p],
                                    "second_name": second_group[0] if second_group else None,
                                    "second_score": scores[second_group[0]] if second_group else None,
                                }
                    groups = rank_groups(scores)
                    pos = 0
                    for group in groups:
                        share = w / len(group)
                        for gp in range(pos, pos + len(group)):
                            for p in group:
                                rank_dist[p][gp] += share
                        pos += len(group)
                    pays = {}
                    for p in players:
                        score_sum[p] += w * scores[p]
                        sv = round(scores[p], 4)
                        score_pairs[p][sv] = score_pairs[p].get(sv, 0.0) + w
                        pay = round(payment_for_rank(ranks[p]), 6)
                        pays[p] = pay
                        cuts_sum[p] += w * pay
                        cuts_pairs[p][pay] = cuts_pairs[p].get(pay, 0.0) + w
                    for m in range(n_matches):
                        if bits[m] == 0:
                            impact_mass_a[m] += w
                            for p in players: impact_cuts_a[m][p] += w * pays[p]
                        else:
                            impact_mass_b[m] += w
                            for p in players: impact_cuts_b[m][p] += w * pays[p]

        total_prob_by_mode[mode] = total_prob
        rank_dist_pct[mode] = {p: [round(100 * v / total_prob, 2) for v in rank_dist[p]] for p in players}

        summary[mode] = {}
        chuleton_exp[mode] = {}
        chuleton_dist[mode] = {}
        for p in players:
            tot = total_prob
            summary[mode][p] = {
                "win_pct": round(100 * win_mass[p] / tot, 2),
                "mean": round(score_sum[p] / tot, 1),
                "p10": wq(score_pairs[p], 0.10, tot),
                "p50": wq(score_pairs[p], 0.50, tot),
                "p90": wq(score_pairs[p], 0.90, tot),
            }
            chuleton_exp[mode][p] = round(cuts_sum[p] / tot, 4)
            chuleton_dist[mode][p] = {
                "min": wq(cuts_pairs[p], 0.0, tot), "p10": wq(cuts_pairs[p], 0.10, tot),
                "q1": wq(cuts_pairs[p], 0.25, tot), "median": wq(cuts_pairs[p], 0.50, tot),
                "q3": wq(cuts_pairs[p], 0.75, tot), "p90": wq(cuts_pairs[p], 0.90, tot),
                "max": wq(cuts_pairs[p], 1.0, tot), "mean": chuleton_exp[mode][p],
            }

        rows = []
        for m in range(n_matches):
            stage, a, b = match_labels[m]
            players_dict = {}
            impact = 0.0
            for p in players:
                ca = round(impact_cuts_a[m][p] / impact_mass_a[m], 4) if impact_mass_a[m] > 0 else 0
                cb = round(impact_cuts_b[m][p] / impact_mass_b[m], 4) if impact_mass_b[m] > 0 else 0
                players_dict[p] = {"cutsA": ca, "cutsB": cb}
                impact += abs(ca - cb)
            rows.append({"stage": stage, "a": a, "b": b,
                         "probA": round(100 * impact_mass_a[m] / total_prob, 1),
                         "probB": round(100 * impact_mass_b[m] / total_prob, 1),
                         "impact": round(impact, 3), "players": players_dict})
        matches_out[mode] = rows
        print(f"{mode} listo", file=sys.stderr)

    players_out = []
    for p in players:
        players_out.append({
            "name": p, "current": current_total[p],
            "ko_min": bracket_min[p], "ko_max": bracket_max[p],
            "total_min": current_total[p] + bracket_min[p],
            "total_max": current_total[p] + bracket_max[p] + 24 + 24,
            "uniform": {"total_avg": summary["uniform"][p]["mean"], "win_pct": summary["uniform"][p]["win_pct"]},
            "weighted": {"mean": summary["weighted"][p]["mean"], "p10": summary["weighted"][p]["p10"],
                         "p50": summary["weighted"][p]["p50"], "p90": summary["weighted"][p]["p90"],
                         "win_pct": summary["weighted"][p]["win_pct"]},
            "dead": dead[p],
            "chuleton": {"uniform": chuleton_exp["uniform"][p], "weighted": chuleton_exp["weighted"][p]},
            "chuleton_box": {"uniform": chuleton_dist["uniform"][p], "weighted": chuleton_dist["weighted"][p]},
            "rank_dist": {"uniform": rank_dist_pct["uniform"][p], "weighted": rank_dist_pct["weighted"][p]},
            "golden_pick": picks[p]["botaoro"], "ball_pick": picks[p]["balonoro"],
        })

    # ---- camino a la victoria: escenario más probable en el que cada
    # jugador queda 1º (solo tiene sentido con probabilidades reales) ----
    def describe_scenario(bits, gname, bname):
        O_winner, Q_winner, S_winner, S_loser, champion, runner, winner34, _ = simulate(bits, "weighted")
        events = []
        for j, i in enumerate(unresolved_idx):
            o = octavos[i]
            events.append({"stage": "Octavos", "a": o["a"], "b": o["b"], "winner": O_winner[i]})
        for k, (ia, ib) in enumerate(qf_pairs):
            a = octavos[ia]["winner"] if octavos[ia]["resolved"] else f"Ganador({octavos[ia]['a']}/{octavos[ia]['b']})"
            b = octavos[ib]["winner"] if octavos[ib]["resolved"] else f"Ganador({octavos[ib]['a']}/{octavos[ib]['b']})"
            events.append({"stage": "Cuartos", "a": a, "b": b, "winner": Q_winner[k]})
        for k, (ia, ib) in enumerate(sf_pairs):
            events.append({"stage": "Semis", "a": Q_winner[ia], "b": Q_winner[ib], "winner": S_winner[k]})
        ia, ib = f_pair
        events.append({"stage": "Final", "a": S_winner[ia], "b": S_winner[ib], "winner": champion})
        ia, ib = tp_pair
        events.append({"stage": "3º-4º puesto", "a": S_loser[ia], "b": S_loser[ib], "winner": winner34})
        return {"events": events, "champion": champion, "runner_up": runner, "tercer_puesto": winner34,
                "golden": gname, "ball": bname}

    camino_out = {}
    for p in players:
        cb = camino_best[p]
        if cb["bits"] is None:
            camino_out[p] = None
            continue
        desc = describe_scenario(cb["bits"], cb["gname"], cb["bname"])
        desc["prob_pct"] = round(100 * cb["w"] / total_prob_by_mode["weighted"], 4)
        desc["score"] = cb["score"]
        desc["tied_with"] = cb["tied_with"]
        desc["second_name"] = cb["second_name"]
        desc["second_score"] = cb["second_score"]
        camino_out[p] = desc

    # ---- bracket con probabilidades (DP exacto, sin enumerar 2^14) ----
    def next_round_dist(pairs, dists, btfn):
        out = []
        for ia, ib in pairs:
            distA, distB = dists[ia], dists[ib]
            result = {}
            for teamA, pA_reach in distA.items():
                win_prob = sum(pB * btfn(teamA, teamB) for teamB, pB in distB.items())
                result[teamA] = pA_reach * win_prob
            for teamB, pB_reach in distB.items():
                win_prob = sum(pA * btfn(teamB, teamA) for teamA, pA in distA.items())
                result[teamB] = pB_reach * win_prob
            out.append(result)
        return out

    def loser_dist(distA, distB, btfn):
        result = {}
        for teamA, pA_reach in distA.items():
            lose_prob = sum(pB * btfn(teamB, teamA) for teamB, pB in distB.items())
            result[teamA] = pA_reach * lose_prob
        for teamB, pB_reach in distB.items():
            lose_prob = sum(pA * btfn(teamA, teamB) for teamA, pA in distA.items())
            result[teamB] = pB_reach * lose_prob
        return result

    O_dist = [({o["winner"]: 1.0} if o["resolved"] else {o["a"]: o["probA"], o["b"]: o["probB"]}) for o in octavos]
    Q_dist = next_round_dist(qf_pairs, O_dist, hybrid_prob)
    S_dist = next_round_dist(sf_pairs, Q_dist, hybrid_prob)
    F_dist = next_round_dist([f_pair], S_dist, hybrid_prob)[0]
    L_dist = [loser_dist(Q_dist[sf_pairs[i][0]], Q_dist[sf_pairs[i][1]], hybrid_prob) for i in range(2)]
    tp_a, tp_b = tp_pair
    T34_dist = next_round_dist([(tp_a, tp_b)], L_dist, hybrid_prob)[0]

    def merge_dists(dists):
        """Une varias distribuciones disjuntas (cada equipo solo puede salir
        de una) sumando sus probabilidades -- para pasar de "2 finalistas
        por separado" a "probabilidad de ser uno de los 2 finalistas"."""
        out = {}
        for d in dists:
            for t, p in d.items():
                out[t] = out.get(t, 0.0) + p
        return out

    # "Clasificarse" al partido de 3º-4º puesto (perder ambas semis) es
    # distinto de "ganarlo". Para la final no hace falta el mismo desglose:
    # "Semifinal 1/2" ya muestra quién gana cada semi, que es exactamente
    # quién llega a la final.
    bronce_dist = merge_dists(L_dist)

    def top_n(dist, n=None):
        """n=None -> todos los equipos que pueden llegar a esa casilla (todo
        el soporte de la distribución), no solo los favoritos."""
        items = sorted(dist.items(), key=lambda kv: -kv[1])
        if n is not None:
            items = items[:n]
        return [{"team": t, "pct": round(p * 100, 1)} for t, p in items]

    bracket = {
        "octavos": [
            ({"a": o["a"], "b": o["b"], "resolved": True, "winner": o["winner"], "score": o["score"],
              "aProbW": 100.0 if o["winner"] == o["a"] else 0.0, "bProbW": 100.0 if o["winner"] == o["b"] else 0.0}
             if o["resolved"] else
             {"a": o["a"], "b": o["b"], "resolved": False,
              "aProbW": round(o["probA"] * 100, 1), "bProbW": round(o["probB"] * 100, 1)})
            for o in octavos
        ],
        "qf": [{"label": f"Cuartos {k+1}", "top": top_n(Q_dist[k]), "from": list(qf_pairs[k])} for k in range(4)],
        "sf": [{"label": f"Semifinal {k+1}", "top": top_n(S_dist[k]), "from": list(sf_pairs[k])} for k in range(2)],
        "campeon": {"top": top_n(F_dist)},
        "clasificados34": {"label": "Clasificados a 3º-4º puesto", "top": top_n(bronce_dist), "fromLosers": True},
        "tercerpuesto": {"label": "3º puesto", "top": top_n(T34_dist), "fromLosers": True},
    }

    # ---- quién va con quién: cuántos picks comparten cada dos jugadores ----
    def pick_list(p):
        pk = picks[p]
        return (pk["cuartofinalista"] + pk["semifinalista"] + pk["finalista"] + pk["tresycuatro"]
                + [pk["campeon"], pk["subcampeon"], pk["tercerpuesto"], pk["botaoro"], pk["balonoro"]])

    pick_lists = {p: pick_list(p) for p in players}
    n_picks = len(pick_lists[players[0]])
    affinity = {}
    for p1 in players:
        affinity[p1] = {}
        for p2 in players:
            if p1 == p2:
                continue
            matches = sum(1 for a, b in zip(pick_lists[p1], pick_lists[p2]) if a == b)
            affinity[p1][p2] = {"matches": matches, "total": n_picks, "pct": round(100 * matches / n_picks, 1)}

    golden = {"candidates": [{"name": n, "pct": round(p * 100, 2)} for n, p in GOLDEN_CANDIDATES],
              "picks": {p: picks[p]["botaoro"] for p in players}}
    ball = {"candidates": [{"name": n, "pct": round(p * 100, 2)} for n, p in GOLDEN_BALL_CANDIDATES],
            "picks": {p: picks[p]["balonoro"] for p in players}}
    elo = [{"team": t, "elo": e} for t, e in sorted(ELO.items(), key=lambda kv: -kv[1])]

    # ---- PRUEBA: simulador interactivo -- picks completos + topología del
    # cuadro para poder recalcular la puntuación en el navegador sin Python ----
    picks_out = {p: picks[p] for p in players}
    topology = {
        "octavos": [
            ({"a": o["a"], "b": o["b"], "resolved": True, "winner": o["winner"]}
             if o["resolved"] else {"a": o["a"], "b": o["b"], "resolved": False,
                                     "favA": o["probA"] >= o["probB"]})
            for o in octavos
        ],
        "qf_pairs": [list(pair) for pair in qf_pairs],
        "sf_pairs": [list(pair) for pair in sf_pairs],
        "f_pair": list(f_pair),
        "tp_pair": list(tp_pair),
    }

    # Para que el simulador pueda calcular el favorito real (no solo en
    # octavos) hace falta poder reproducir hybrid_prob() en el navegador.
    known_matchups_out = []
    for key, odds in KNOWN_MATCHUPS.items():
        a, b = list(key)
        known_matchups_out.append({"a": a, "b": b, "probA": round(odds[a] / (odds[a] + odds[b]), 4)})

    n_unresolved = len(unresolved_idx)
    payload = {
        "generated_from_last_updated": porra.get("last_updated"),
        "n_remaining_matches": n_bits,
        "n_combinations": len(all_bits),
        "players": players_out,
        "matches_uniform": matches_out["uniform"],
        "matches_weighted": matches_out["weighted"],
        "bracket": bracket,
        "golden": golden,
        "ball": ball,
        "elo": elo,
        "camino": camino_out,
        "affinity": affinity,
        "picks": picks_out,
        "topology": topology,
        "current_total": current_total,
        "known_matchups": known_matchups_out,
    }

    out_path = ROOT / "predictions_data.js"
    out_path.write_text("const PREDICTIONS_DATA = " + json.dumps(payload, ensure_ascii=False) + ";\n", encoding="utf-8")
    print(f"Escrito {out_path} ({n_unresolved} octavos sin jugar, {n_bits} bits, {len(all_bits)} combinaciones)", file=sys.stderr)

if __name__ == "__main__":
    main()
